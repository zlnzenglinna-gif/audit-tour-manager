# services/kostenerstellung_service.py

from copy import copy
from datetime import timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


# ============================================================
# COPY CELL STYLE
# ============================================================

def copy_cell_style(
    source_cell,
    target_cell,
):

    if source_cell.has_style:
        target_cell._style = copy(
            source_cell._style
        )

    target_cell.font = copy(
        source_cell.font
    )

    target_cell.fill = copy(
        source_cell.fill
    )

    target_cell.border = copy(
        source_cell.border
    )

    target_cell.alignment = copy(
        source_cell.alignment
    )

    target_cell.protection = copy(
        source_cell.protection
    )

    target_cell.number_format = (
        source_cell.number_format
    )


# ============================================================
# COPY ROW STYLE
# ============================================================

def copy_row_style(
    worksheet,
    source_row,
    target_row,
):

    for column in range(
        1,
        10,
    ):

        source_cell = worksheet.cell(
            row=source_row,
            column=column,
        )

        target_cell = worksheet.cell(
            row=target_row,
            column=column,
        )

        if isinstance(
            target_cell,
            MergedCell,
        ):
            continue

        copy_cell_style(
            source_cell,
            target_cell,
        )

    source_height = (
        worksheet.row_dimensions[
            source_row
        ].height
    )

    if source_height is not None:

        worksheet.row_dimensions[
            target_row
        ].height = source_height


# ============================================================
# CLEAR CELL SAFELY
# ============================================================

def clear_cell(
    worksheet,
    row,
    column,
):

    cell = worksheet.cell(
        row=row,
        column=column,
    )

    if isinstance(
        cell,
        MergedCell,
    ):
        return

    cell.value = None


# ============================================================
# FORMAT DATE PERIOD
# ============================================================

def format_date_period(
    start_date,
    end_date,
):

    if (
        start_date is None
        or end_date is None
    ):

        return ""

    return (
        f"{start_date.strftime('%d.%m.')}"
        f"-{end_date.strftime('%d.%m.%Y')}"
    )


# ============================================================
# BUILD HOTEL PERIODS
# ============================================================

def build_hotel_periods(
    hotels,
    interpreter_end_date,
):
    """
    Determine hotel Check-in / Check-out.

    Check-in:
        Hotel row date.

    Check-out:
        Next hotel check-in date.

    For the last hotel:
        interpreter end date.

    If check-out <= check-in:
        check-out = check-in + 1 day.
    """

    valid_hotels = [
        hotel
        for hotel in hotels
        if hotel.get("date") is not None
    ]

    valid_hotels.sort(
        key=lambda hotel: hotel["date"]
    )

    result = []

    for index, hotel in enumerate(
        valid_hotels
    ):

        check_in = hotel["date"]

        if (
            index + 1
            < len(valid_hotels)
        ):

            check_out = (
                valid_hotels[
                    index + 1
                ]["date"]
            )

        else:

            check_out = (
                interpreter_end_date
            )

        if (
            check_out is None
            or check_out <= check_in
        ):

            check_out = (
                check_in
                + timedelta(days=1)
            )

        result.append(
            {
                **hotel,

                "check_in":
                    check_in,

                "check_out":
                    check_out,

                "date_text":
                    format_date_period(
                        check_in,
                        check_out,
                    ),
            }
        )

    return result


# ============================================================
# PREPARE COST ROW
# ============================================================

def prepare_cost_row(
    worksheet,
    row_number,
    source_style_row,
):

    # Copy visual style
    copy_row_style(
        worksheet=worksheet,
        source_row=source_style_row,
        target_row=row_number,
    )

    # Make sure A:C is merged like the template.
    merged_range = (
        f"A{row_number}:C{row_number}"
    )

    already_merged = any(
        str(merged) == merged_range
        for merged
        in worksheet.merged_cells.ranges
    )

    if not already_merged:

        worksheet.merge_cells(
            merged_range
        )


# ============================================================
# WRITE STANDARD COST ROW
# ============================================================

def write_cost_row(
    worksheet,
    row_number,
    description,
    date_text,
    amount_or_formula,
    comment="",
):

    worksheet.cell(
        row=row_number,
        column=1,
    ).value = description

    worksheet.cell(
        row=row_number,
        column=4,
    ).value = date_text

    worksheet.cell(
        row=row_number,
        column=5,
    ).value = amount_or_formula

    worksheet.cell(
        row=row_number,
        column=6,
    ).value = "EUR"

    worksheet.cell(
        row=row_number,
        column=7,
    ).value = 1.0

    worksheet.cell(
        row=row_number,
        column=8,
    ).value = (
        f"=E{row_number}/G{row_number}"
    )

    worksheet.cell(
        row=row_number,
        column=9,
    ).value = comment


# ============================================================
# EXPORT KOSTENERSTELLUNG
# ============================================================

def export_kostenerstellung_excel(
    template_excel_bytes,
    interpreter_name,
    start_date,
    end_date,
    interpreter_cost,
    meal_excel_formula=None,
    sonstiges_value=0.0,
):

    print("### NEW KOSTENERSTELLUNG EXPORT IS RUNNING ###")

    # ========================================================
    # LOAD TEMPLATE
    # ========================================================

    workbook = load_workbook(
        BytesIO(
            template_excel_bytes
        )
    )

    worksheet = workbook.active

    # ========================================================
    # BASIC INFORMATION
    # ========================================================

    worksheet["D6"] = (
        interpreter_name
    )

    worksheet["D7"] = (
        start_date
    )

    worksheet["D8"] = (
        end_date
    )

    # ========================================================
    # DATE FORMAT
    # ========================================================

    worksheet["D7"].number_format = (
        "DD.MM.YYYY"
    )

    worksheet["D8"].number_format = (
        "DD.MM.YYYY"
    )


    # ========================================================
    # GET CALCULATED COST DATA
    # ========================================================

    hotels = interpreter_cost.get(
        "hotels",
        [],
    )

    fuel_formula = (
        interpreter_cost.get(
            "fuel_excel_formula",
            "=0",
        )
        or "=0"
    )

    if meal_excel_formula is None:

        meal_excel_formula = (
            interpreter_cost.get(
                "meal_excel_formula",
                "=0",
            )
            or "=0"
        )


    # ========================================================
    # HOTEL PERIODS
    #
    # IMPORTANT:
    # Build periods from ALL hotels first.
    # This ensures a bezahlt hotel can still determine the
    # checkout date of the previous vor-Ort hotel.
    # ========================================================

    all_hotel_periods = (
        build_hotel_periods(
            hotels=hotels,
            interpreter_end_date=end_date,
        )
    )


    # Only "vor Ort" is transferred to interpreter
    hotel_rows = [
        hotel
        for hotel
        in all_hotel_periods
        if str(
            hotel.get(
                "payment_status",
                ""
            )
        ).strip().lower()
        == "vor ort"
    ]


    # ========================================================
    # WRITE COST TABLE
    # ========================================================

    current_row = 16


    # --------------------------------------------------------
    # HOTELS
    # --------------------------------------------------------

    for hotel in hotel_rows:

        # Template hotel style = row 16
        prepare_cost_row(
            worksheet=worksheet,
            row_number=current_row,
            source_style_row=16,
        )

        write_cost_row(
            worksheet=worksheet,
            row_number=current_row,
            description=(
                hotel.get(
                    "hotel",
                    ""
                )
            ),
            date_text=(
                hotel[
                    "date_text"
                ]
            ),
            amount_or_formula=(
                hotel.get(
                    "price",
                    0.0,
                )
            ),
            comment="vor Ort",
        )

        current_row += 1


    # --------------------------------------------------------
    # TANKEN
    # --------------------------------------------------------

    prepare_cost_row(
        worksheet=worksheet,
        row_number=current_row,
        source_style_row=21,
    )

    write_cost_row(
        worksheet=worksheet,
        row_number=current_row,
        description="Tanken",
        date_text=(
            format_date_period(
                start_date,
                end_date,
            )
        ),
        amount_or_formula=(
            fuel_formula
        ),
    )

    current_row += 1


    # --------------------------------------------------------
    # ESSEN
    # --------------------------------------------------------

    prepare_cost_row(
        worksheet=worksheet,
        row_number=current_row,
        source_style_row=22,
    )

    write_cost_row(
        worksheet=worksheet,
        row_number=current_row,
        description=(
            "Essen, Getränke und Snacks"
        ),
        date_text=(
            format_date_period(
                start_date,
                end_date,
            )
        ),
        amount_or_formula=(
            meal_excel_formula
        ),
    )

    current_row += 1

    # --------------------------------------------------------
    # SONSTIGES
    # --------------------------------------------------------

    prepare_cost_row(
        worksheet=worksheet,
        row_number=current_row,
        source_style_row=23,
    )

    sonstiges_row = current_row

    last_cost_row_before_sonstiges = (
        sonstiges_row - 1
    )

    sonstiges_formula = (
        f"=CEILING("
        f"SUM(H16:H{last_cost_row_before_sonstiges}),"
        f"10"
        f")"
        f"-SUM(H16:H{last_cost_row_before_sonstiges})"
    )

    write_cost_row(
        worksheet=worksheet,
        row_number=sonstiges_row,
        description="Sonstiges",
        date_text=(
            format_date_period(
                start_date,
                end_date,
            )
        ),
        amount_or_formula=sonstiges_formula,
    )

    current_row += 1


    # --------------------------------------------------------
    # SUMME
    # --------------------------------------------------------

    prepare_cost_row(
        worksheet=worksheet,
        row_number=current_row,
        source_style_row=24,
    )

    worksheet.cell(
        row=current_row,
        column=1,
    ).value = "Summe"

    worksheet.cell(
        row=current_row,
        column=8,
    ).value = (
        f"=SUM(H16:H{current_row - 1})"
    )


    # ========================================================
    # CLEAR ANY REMAINING OLD TEMPLATE ROWS
    # ========================================================

    for row_number in range(
        current_row + 1,
        25,
    ):

        clear_cell(
            worksheet,
            row_number,
            1,
        )

        for column in range(
            4,
            10,
        ):

            clear_cell(
                worksheet,
                row_number,
                column,
            )


    # ========================================================
    # OUTPUT
    # ========================================================

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output