from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook


COMPANY_SUFFIXES = [
    "gmbh",
    "mbh",
    "ag",
    "kg",
    "kgaa",
    "s.r.o",
    "sro",
    "d.o.o",
    "doo",
    "kft",
    "ltd",
    "limited",
    "inc",
    "llc",
    "sa",
]


def normalize_location_name(value: object) -> str:
    """清理公司名称，方便后续匹配Hotelübersicht。"""
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = re.sub(r"[.,;:()/_-]+", " ", text)
    text = re.sub(r"\s+", " ", text)

    words = [
        word
        for word in text.split()
        if word not in COMPANY_SUFFIXES
    ]

    return " ".join(words).strip()


def format_excel_date(value: object) -> str:
    """把Excel日期统一转换成DD.MM.YYYY。"""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")

    if value is None:
        return ""

    return str(value).strip()


def get_fill_color(cell) -> str:
    """读取单元格背景色。"""
    fill = cell.fill

    if fill.fill_type is None:
        return ""

    color = fill.fgColor

    if color.type == "rgb":
        return (color.rgb or "").upper()

    if color.type == "indexed":
        return f"INDEXED:{color.indexed}"

    if color.type == "theme":
        return f"THEME:{color.theme}"

    return ""


def is_yellow_cell(cell) -> bool:
    """判断单元格是否为黄色。"""
    color = get_fill_color(cell)

    return color in {
        "FFFFFF00",
        "00FFFF00",
    }


def read_yellow_factories(
    file_source: str | Path | BinaryIO,
) -> pd.DataFrame:
    """
    只读取Tourplan中D列标黄色的工厂。

    同时读取：
    A列：日期
    D列：工厂名称
    E列：地址
    Excel原始行号
    """

    workbook = load_workbook(
        file_source,
        data_only=False,
    )

    worksheet = workbook.active

    records: list[dict[str, object]] = []

    for row_number in range(
        1,
        worksheet.max_row + 1,
    ):
        factory_cell = worksheet.cell(
            row=row_number,
            column=4,
        )

        if not is_yellow_cell(factory_cell):
            continue

        factory_value = factory_cell.value

        if factory_value is None:
            continue

        factory_name = str(factory_value).strip()

        if not factory_name:
            continue

        date_value = worksheet.cell(
            row=row_number,
            column=1,
        ).value

        address_value = worksheet.cell(
            row=row_number,
            column=5,
        ).value

        address = (
            ""
            if address_value is None
            else str(address_value).strip()
        )

        records.append(
            {
                "Excel Row": row_number,
                "Date": format_excel_date(
                    date_value
                ),
                "Factory": factory_name,
                "Address": address,
                "Normalized Factory": (
                    normalize_location_name(
                        factory_name
                    )
                ),
            }
        )

    return pd.DataFrame(records)

def read_tourplan_timeline(
    file_source,
) -> pd.DataFrame:
    """
    Read the original Tourplan timeline.

    Excel columns:
    A = Date
    B = Day
    C = Travel plan
    D = Factory/Hotel
    E = Address

    Only rows with a real date in column A are included.
    """

    workbook = load_workbook(
        file_source,
        data_only=False,
    )

    worksheet = workbook.active

    records = []

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):

        date_value = worksheet.cell(
            row=row_number,
            column=1,
        ).value

        day_value = worksheet.cell(
            row=row_number,
            column=2,
        ).value

        travel_plan = worksheet.cell(
            row=row_number,
            column=3,
        ).value

        factory_hotel = worksheet.cell(
            row=row_number,
            column=4,
        ).value

        address = worksheet.cell(
            row=row_number,
            column=5,
        ).value


        # ==================================================
        # 1. Nur echte Datumszeilen behalten
        # ==================================================

        if date_value in (None, ""):
            continue

        # Excel kann Datum entweder als datetime/date
        # oder als Text enthalten.
        date_text = format_excel_date(
            date_value
        )

        # Wenn format_excel_date keinen echten
        # Datumswert erkannt hat, Zeile überspringen.
        if not re.match(
            r"^\d{2}\.\d{2}\.\d{4}$",
            date_text,
        ):
            continue


        # ==================================================
        # 2. Originalwerte übernehmen
        # ==================================================

        day_text = (
            str(day_value).strip()
            if day_value not in (None, "")
            else ""
        )

        travel_plan_text = (
            str(travel_plan).strip()
            if travel_plan not in (None, "")
            else ""
        )

        factory_hotel_text = (
            str(factory_hotel).strip()
            if factory_hotel not in (None, "")
            else ""
        )

        address_text = (
            str(address).strip()
            if address not in (None, "")
            else ""
        )


        # ==================================================
        # 3. Nur prüfen: Ist D gelb?
        # ==================================================

        factory_cell = worksheet.cell(
            row=row_number,
            column=4,
        )

        fill_color = (
            factory_cell.fill.fgColor.rgb
        )

        is_yellow_factory = False

        if fill_color:

            color_text = str(
                fill_color
            ).upper()

            if color_text.endswith(
                "FFFF00"
            ):
                is_yellow_factory = True


        # ==================================================
        # 4. Timeline-Zeile speichern
        # ==================================================

        records.append(
            {
                "Date": date_text,
                "Day": day_text,
                "Travel plan": travel_plan_text,
                "Factory/Hotel": factory_hotel_text,
                "Address": address_text,

                # Intern:
                "Is Factory": is_yellow_factory,
                "Excel Row": row_number,
            }
        )


    return pd.DataFrame(
        records
    )