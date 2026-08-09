from copy import copy
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ============================================================
# COUNTRY MAPPING
# ============================================================

COUNTRY_MAPPING = {
    "germany": "Deutschland",
    "deutschland": "Deutschland",

    "austria": "Österreich",
    "österreich": "Österreich",

    "slovakia": "Slovakia",
    "slovakei": "Slovakia",

    "serbia": "Serbien",
    "serbien": "Serbien",

    "hungary": "Ungarn",
    "ungarn": "Ungarn",

    "poland": "Polen",
    "polen": "Polen",

    "czech republic": "Tschechien",
    "czechia": "Tschechien",
    "tschechien": "Tschechien",

    "romania": "Rumänien",
    "rumänien": "Rumänien",

    "croatia": "Kroatien",
    "kroatien": "Kroatien",

    "slovenia": "Slowenien",
    "slowenien": "Slowenien",

    "france": "Frankreich",
    "frankreich": "Frankreich",

    "italy": "Italien",
    "italien": "Italien",

    "spain": "Spanien",
    "spanien": "Spanien",

    "belgium": "Belgien",
    "belgien": "Belgien",

    "netherlands": "Niederlande",
    "niederlande": "Niederlande",

    "switzerland": "Schweiz",
    "schweiz": "Schweiz",
}


# ============================================================
# COUNTRY DETECTION
# ============================================================

def detect_country_from_address(address):
    """
    Detect country from an address.

    Existing Excel Country values are preserved.
    This is mainly used for empty Country cells
    and manually added rows.
    """

    if not address:
        return ""

    text = str(address).strip().lower()

    address_parts = [
        part.strip()
        for part in text.split(",")
        if part.strip()
    ]

    # First try the last address part
    if address_parts:

        last_part = address_parts[-1]

        for keyword, country in COUNTRY_MAPPING.items():

            if keyword in last_part:
                return country

    # Fallback: search complete address
    for keyword, country in COUNTRY_MAPPING.items():

        if keyword in text:
            return country

    return ""


# ============================================================
# WEEKDAY
# ============================================================

def calculate_day(date_text):

    day_mapping = {
        0: "Mo",
        1: "Di",
        2: "Mi",
        3: "Do",
        4: "Fr",
        5: "Sa",
        6: "So",
    }

    try:

        parsed_date = datetime.strptime(
            str(date_text).strip(),
            "%d.%m.%Y",
        )

        return day_mapping[
            parsed_date.weekday()
        ]

    except (ValueError, TypeError):

        return ""


# ============================================================
# DATE CONVERSION
# ============================================================

def convert_date_for_excel(date_text):
    """
    Convert DD.MM.YYYY into datetime.
    """

    if not date_text:
        return None

    try:

        return datetime.strptime(
            str(date_text).strip(),
            "%d.%m.%Y",
        )

    except ValueError:

        return date_text


# ============================================================
# HOTEL PRICE TEXT
# ============================================================

def build_hotel_price_text(
    price,
    payment_status,
):
    """
    Examples:

    180 + vor Ort
    -> 180 EUR (vor Ort)

    165 + bezahlt
    -> 165 EUR (bezahlt)
    """

    if price is None:
        return None

    try:

        numeric_price = float(
            price
        )

    except (
        TypeError,
        ValueError,
    ):

        return None


    if numeric_price.is_integer():

        price_text = str(
            int(
                numeric_price
            )
        )

    else:

        price_text = (
            f"{numeric_price:.2f}"
            .rstrip("0")
            .rstrip(".")
        )


    if payment_status:

        return (
            f"{price_text} EUR "
            f"({payment_status})"
        )

    return (
        f"{price_text} EUR"
    )


# ============================================================
# BREAKFAST
# ============================================================

def convert_breakfast_value(value):

    if not value:
        return None

    normalized = str(
        value
    ).strip().lower()


    if normalized == "inklusive":
        return "yes"

    if normalized == "nicht inklusive":
        return "no"

    return value


# ============================================================
# COPY ROW STYLE
# ============================================================

def copy_row_style(
    worksheet,
    source_row,
    target_row,
):
    """
    Copy formatting from one row to another.

    Values are not copied.
    """

    for column in range(
        1,
        worksheet.max_column + 1,
    ):

        source_cell = worksheet.cell(
            row=source_row,
            column=column,
        )

        target_cell = worksheet.cell(
            row=target_row,
            column=column,
        )


        if source_cell.has_style:

            target_cell._style = copy(
                source_cell._style
            )


        target_cell.font = copy(
            source_cell.font
        )

        target_cell.fill = copy(
            source_cell.fill
        )

        target_cell.border = copy(
            source_cell.border
        )

        target_cell.alignment = copy(
            source_cell.alignment
        )

        target_cell.protection = copy(
            source_cell.protection
        )

        target_cell.number_format = (
            source_cell.number_format
        )


    source_height = (
        worksheet.row_dimensions[
            source_row
        ].height
    )


    if source_height is not None:

        worksheet.row_dimensions[
            target_row
        ].height = source_height


# ============================================================
# CLEAR MANUAL ROW FACTORY COLOR
# ============================================================

def clear_manual_row_factory_fill(
    worksheet,
    row_number,
):
    """
    Manual Add Row is NOT treated as a factory.

    Column D is therefore forced to white so it does not
    inherit the yellow Werk formatting from the anchor row.
    """

    factory_hotel_cell = worksheet.cell(
        row=row_number,
        column=4,
    )

    factory_hotel_cell.fill = PatternFill(
        fill_type="solid",
        fgColor="FFFFFF",
        bgColor="FFFFFF",
    )


# ============================================================
# CLEAR MANUAL ROW NON-FACTORY BACKGROUND
# ============================================================

def clear_manual_row_extra_factory_style(
    worksheet,
    row_number,
):
    """
    Make sure the manually inserted row does not look
    like a yellow factory row.

    At minimum D must be white.
    """

    clear_manual_row_factory_fill(
        worksheet=worksheet,
        row_number=row_number,
    )


# ============================================================
# WRITE ONE TOUR ROW
# ============================================================

def write_tour_row(
    worksheet,
    excel_row,
    date_value,
    travel_plan,
    factory_hotel,
    address,
    distance=None,
    hotel_price=None,
    payment_status="",
    breakfast="",
    preserve_existing_country=True,
):
    """
    Column mapping:

    A = Date
    B = Day
    C = Travel plan
    D = Factory/Hotel
    E = Address
    F = Country
    G = Comment       untouched
    H = Consultant    untouched
    I = Distance
    J = Hotel price/status
    K = Breakfast
    L+ = untouched
    """

    # ========================================================
    # A - DATE
    # ========================================================

    converted_date = (
        convert_date_for_excel(
            date_value
        )
    )

    if converted_date is not None:

        worksheet.cell(
            row=excel_row,
            column=1,
        ).value = converted_date


    # ========================================================
    # B - DAY
    # ========================================================

    day_value = calculate_day(
        date_value
    )

    if day_value:

        worksheet.cell(
            row=excel_row,
            column=2,
        ).value = day_value


    # ========================================================
    # C - TRAVEL PLAN
    # ========================================================

    worksheet.cell(
        row=excel_row,
        column=3,
    ).value = (
        travel_plan or ""
    )


    # ========================================================
    # D - FACTORY / HOTEL
    # ========================================================

    worksheet.cell(
        row=excel_row,
        column=4,
    ).value = (
        factory_hotel or ""
    )


    # ========================================================
    # E - ADDRESS
    # ========================================================

    worksheet.cell(
        row=excel_row,
        column=5,
    ).value = (
        address or ""
    )


    # ========================================================
    # F - COUNTRY
    # ========================================================

    country_cell = worksheet.cell(
        row=excel_row,
        column=6,
    )

    existing_country = (
        country_cell.value
    )


    if (
        not preserve_existing_country
        or not existing_country
    ):

        detected_country = (
            detect_country_from_address(
                address
            )
        )


        if detected_country:

            country_cell.value = (
                detected_country
            )


    # ========================================================
    # I - DISTANCE
    # ========================================================

    if distance is not None:

        worksheet.cell(
            row=excel_row,
            column=9,
        ).value = distance


    # ========================================================
    # J - HOTEL PRICE / STATUS
    # ========================================================

    hotel_price_text = (
        build_hotel_price_text(
            hotel_price,
            payment_status,
        )
    )


    if hotel_price_text is not None:

        worksheet.cell(
            row=excel_row,
            column=10,
        ).value = (
            hotel_price_text
        )


    # ========================================================
    # K - BREAKFAST
    # ========================================================

    breakfast_value = (
        convert_breakfast_value(
            breakfast
        )
    )


    if breakfast_value is not None:

        worksheet.cell(
            row=excel_row,
            column=11,
        ).value = (
            breakfast_value
        )


# ============================================================
# EXPORT TOURPLAN
# ============================================================

def export_tourplan_excel(
    original_excel_bytes,
    timeline_df,
    session_state,
):
    """
    Export the edited Tourplan based on the original workbook.

    Original workbook information and formatting are preserved
    as much as possible.

    Manual rows are inserted directly after the selected
    original Excel row.
    """

    workbook = load_workbook(
        BytesIO(
            original_excel_bytes
        )
    )

    worksheet = workbook.active


    # ========================================================
    # 1. UPDATE ORIGINAL EXCEL ROWS
    # ========================================================

    for _, row in timeline_df.iterrows():

        excel_row = int(
            row["Excel Row"]
        )

        row_key = (
            f"timeline_{excel_row}"
        )


        date_value = str(
            row["Date"]
        ).strip()


        travel_plan = (
            session_state.get(
                f"{row_key}_travel_plan",
                str(
                    row["Travel plan"]
                ).strip(),
            )
        )


        factory_hotel = (
            session_state.get(
                f"{row_key}_factory_hotel",
                str(
                    row["Factory/Hotel"]
                ).strip(),
            )
        )


        address = (
            session_state.get(
                f"{row_key}_address",
                str(
                    row["Address"]
                ).strip(),
            )
        )


        distance = (
            session_state.get(
                f"{row_key}_distance",
                None,
            )
        )


        hotel_price = (
            session_state.get(
                f"{row_key}_hotel_price",
                None,
            )
        )


        payment_status = (
            session_state.get(
                f"{row_key}_payment_status",
                "",
            )
        )


        breakfast = (
            session_state.get(
                f"{row_key}_breakfast",
                "",
            )
        )


        write_tour_row(
            worksheet=worksheet,
            excel_row=excel_row,
            date_value=date_value,
            travel_plan=travel_plan,
            factory_hotel=factory_hotel,
            address=address,
            distance=distance,
            hotel_price=hotel_price,
            payment_status=payment_status,
            breakfast=breakfast,
            preserve_existing_country=True,
        )


    # ========================================================
    # 2. INSERT MANUAL ROWS
    #
    # Work bottom-to-top.
    # This keeps original row references stable.
    # ========================================================

    original_excel_rows = sorted(
        [
            int(value)
            for value
            in timeline_df[
                "Excel Row"
            ].tolist()
        ],
        reverse=True,
    )


    for anchor_excel_row in (
        original_excel_rows
    ):

        manual_rows_here = [
            manual_row
            for manual_row
            in session_state.added_rows
            if (
                manual_row[
                    "after_excel_row"
                ]
                == anchor_excel_row
            )
        ]


        if not manual_rows_here:
            continue


        # ----------------------------------------------------
        # Insert required number of rows
        # ----------------------------------------------------

        worksheet.insert_rows(
            anchor_excel_row + 1,
            amount=len(
                manual_rows_here
            ),
        )


        # ----------------------------------------------------
        # Write inserted rows
        # ----------------------------------------------------

        for offset, manual_row in enumerate(
            manual_rows_here,
            start=1,
        ):

            target_row = (
                anchor_excel_row
                + offset
            )


            row_id = (
                manual_row["id"]
            )


            # =================================================
            # COPY FORMAT FROM PREVIOUS ORIGINAL ROW
            # =================================================

            copy_row_style(
                worksheet=worksheet,
                source_row=anchor_excel_row,
                target_row=target_row,
            )


            # =================================================
            # IMPORTANT:
            # MANUAL ROW IS NOT A FACTORY
            # =================================================

            clear_manual_row_extra_factory_style(
                worksheet=worksheet,
                row_number=target_row,
            )


            # =================================================
            # READ MANUAL ROW DATA
            # =================================================

            date_value = (
                session_state.get(
                    f"manual_{row_id}_date",
                    "",
                )
            )


            travel_plan = (
                session_state.get(
                    f"manual_{row_id}_travel_plan",
                    "",
                )
            )


            factory_hotel = (
                session_state.get(
                    f"manual_{row_id}_factory_hotel",
                    "",
                )
            )


            address = (
                session_state.get(
                    f"manual_{row_id}_address",
                    "",
                )
            )


            distance = (
                session_state.get(
                    f"manual_{row_id}_distance",
                    None,
                )
            )


            hotel_price = (
                session_state.get(
                    f"manual_{row_id}_hotel_price",
                    None,
                )
            )


            payment_status = (
                session_state.get(
                    f"manual_{row_id}_payment_status",
                    "",
                )
            )


            breakfast = (
                session_state.get(
                    f"manual_{row_id}_breakfast",
                    "",
                )
            )


            # =================================================
            # WRITE MANUAL ROW
            # =================================================

            write_tour_row(
                worksheet=worksheet,
                excel_row=target_row,
                date_value=date_value,
                travel_plan=travel_plan,
                factory_hotel=factory_hotel,
                address=address,
                distance=distance,
                hotel_price=hotel_price,
                payment_status=payment_status,
                breakfast=breakfast,
                preserve_existing_country=False,
            )


            # =================================================
            # FORCE WHITE AGAIN AFTER WRITING
            #
            # Extra safety: even if some formatting changes
            # during write, D must remain white.
            # =================================================

            clear_manual_row_factory_fill(
                worksheet=worksheet,
                row_number=target_row,
            )


    # ========================================================
    # 3. SAVE TO MEMORY
    # ========================================================

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output